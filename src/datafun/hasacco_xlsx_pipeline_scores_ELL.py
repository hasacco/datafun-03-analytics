"""case_xlsx_pipeline.py - XLSX ETVL pipeline.

Author: Denise Case, Hannah Sacco
Date: 2026-04

  Practice key Python skills related to:
    - ETVL pipeline structure (Extract, Transform, Verify, Load)
    - reading Excel files using the openpyxl package
    - accessing cells by column letter
    - keyword-only function arguments
    - runtime type checking with isinstance()
    - counting word occurrences across strings
    - writing results to a text file

  Paths (relative to repo root):

    INPUT FILE:  data/raw/grade_8_ELL_math_scores_TX.xlsx
    OUTPUT FILE: data/processed/grade_8_ELL_math_scores_TX_stats.txt

  Terminal command to run this file from the root project folder:

    uv run python -m datafun.hasacco_xlsx_pipeline_scores_ELL

OBS:
  Don't edit this file - it should remain a working example.
  Copy it, rename it, and modify your copy.

  This is a copy of the instructor's file and has been edited.
"""

# === DECLARE IMPORTS (BRING IN FREE CODE) ===

import statistics
from pathlib import Path
from typing import Any, cast

import matplotlib.pyplot as plt

# openpyxl is an external package - it must be listed in pyproject.toml dependencies.
# OBS: If you see "import openpyxl could not be resolved", open pyproject.toml,
#      find the dependencies section, and confirm openpyxl is listed there.
#      Then run: uv sync --extra dev --extra docs --upgrade
import openpyxl
from openpyxl.cell.cell import Cell

# === SKILL: READING AN EXCEL FILE WITH openpyxl ===

# openpyxl.load_workbook() opens an Excel file and returns a Workbook object.
# workbook.active returns the first (active) worksheet.
# sheet["A"] returns all cells in column A as a tuple.
# Each cell has a .value attribute containing the cell's contents.
# Cell values can be str, int, float, None, or other types.
# Use isinstance() to check the type before using the value.
# cast() tells the type checker what type to expect - it has no effect at runtime.


# === E: EXTRACT ===


def extract_xlsx_column_tuples(
    *, file_path: Path, column_letter: str, column_letter_key: str
) -> tuple[list[float], list[float]]:
    """E: Read an Excel file and extract numeric values from a column.

    Arguments:
        file_path: Path to input XLSX file.
        column_letter: Letter of the column to extract (e.g., 'A').
        column_letter_key: Letter of the column containing the key (e.g., 'C').

    Returns:
        List of non-empty numeric values from the specified column.
    """
    # Handle known possible error: no file at the path provided.
    if not file_path.exists():
        raise FileNotFoundError(f"Missing input file: {file_path}")

    workbook = openpyxl.load_workbook(file_path)
    # active returns the first worksheet - the one visible when the file opens.
    sheet = workbook.active

    ELL_values: list[float] = []
    Non_ELL_values: list[float] = []

    for cell_key in sheet[column_letter_key]:
        # print(cell_key.value) #Debugging
        target_row = cell_key.row
        if cell_key.value == 'ELL':
            cell = sheet[f"{column_letter}{target_row}"]
            # cast() narrows the type for the type checker - no runtime effect.
            cell = cast(Cell, cell)
            value = cell.value
            # Only keep non-empty numeric values.
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                ELL_values.append(float(value))
        else:
            cell = sheet[f"{column_letter}{target_row}"]
            # cast() narrows the type for the type checker - no runtime effect.
            cell = cast(Cell, cell)
            value = cell.value
            # Only keep non-empty numeric values.
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                Non_ELL_values.append(float(value))

    return ELL_values, Non_ELL_values


def extract_xlsx_column_values(*, file_path: Path, column_letter: str) -> list[float]:
    """E: Read an Excel file and extract numeric values from a column.

    Arguments:
        file_path: Path to input XLSX file.
        column_letter: Letter of the column to extract (e.g., 'A').

    Returns:
        List of non-empty numeric values from the specified column.
    """
    # Handle known possible error: no file at the path provided.
    if not file_path.exists():
        raise FileNotFoundError(f"Missing input file: {file_path}")

    workbook = openpyxl.load_workbook(file_path)
    # active returns the first worksheet - the one visible when the file opens.
    sheet = workbook.active

    values: list[float] = []

    for cell in sheet[column_letter]:
        # cast() narrows the type for the type checker - no runtime effect.
        cell = cast(Cell, cell)
        value = cell.value
        # Only keep non-empty numeric values.
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            values.append(float(value))

    return values


# === T: TRANSFORM ===

# The statistics module provides mean() and stdev().
# stdev() requires at least two values - guard against a single-value list.


def transform_scores_to_stats(*, scores: list[float]) -> dict[str, float]:
    """T: Calculate basic statistics for a list of floats.

    Arguments:
        scores: List of float values.

    Returns:
        Dictionary with keys: count, min, max, mean, stdev.
    """
    if not scores:
        raise ValueError("No numeric values found for analysis.")

    return {
        "count": float(len(scores)),
        "min": min(scores),
        "max": max(scores),
        "mean": statistics.mean(scores),
        "median": statistics.median(scores),
        # stdev() requires at least 2 values; return 0.0 for a single value.
        "stdev": statistics.stdev(scores) if len(scores) > 1 else 0.0,
    }


def plot_scores(
    ELL_scores: list[float],
    Non_ELL_scores: list[float],
    years: list[float],
    out_path: Path,
) -> None:
    """T: Create a line graph of the scores and save it as an image.

    Arguments:
        scores: List of float values to plot.
        years: List of years corresponding to the scores.
        out_path: Path to save the line graph image.

    Returns:
        None
    """

    years = [int(year) for year in years]
    year_list = list(dict.fromkeys(years))  # Remove duplicates while preserving order
    # print(scores) #Debugging
    # print(year_list) #Debugging

    plt.figure(figsize=(10, 6))
    plt.plot(year_list, ELL_scores, label='ELL', marker='o', linestyle='-', color='red')
    plt.plot(
        year_list,
        Non_ELL_scores,
        label='Not ELL',
        marker='o',
        linestyle='-',
        color='blue',
    )
    plt.legend()
    plt.title('Trend of Texas Grade 8 NAEP Math Scores by ELL Status (1990 - 2024)')
    plt.xlabel('Year')
    plt.ylabel('Score')
    plt.grid(axis='y', alpha=0.75)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    plt.savefig(out_path)
    plt.close()


# === V: VERIFY ===

# Verification catches problems between Transform and Load.
# It is cheaper to detect a bad result before writing it to disk.
# Use raise to signal an error the caller must handle.


def verify_stats(*, stats: dict[str, float]) -> None:
    """V: Sanity-check the stats dictionary.

    Arguments:
        stats: Dictionary with statistics to verify.

    Returns:
        None
    """
    required = {"count", "min", "max", "mean", "median", "stdev"}
    missing = required - set(stats.keys())
    # Handle known possible error: missing required keys.
    if missing:
        raise KeyError(f"Missing stats keys: {sorted(missing)}")

    # Handle known possible error: count must be positive.
    if stats["count"] <= 0:
        raise ValueError("Count must be positive.")

    # Handle known possible error: min cannot be greater than max.
    if stats["min"] > stats["max"]:
        raise ValueError("Min cannot be greater than max.")


# === L: LOAD ===

# Path.open("w") creates or overwrites a file.
# Always create parent directories before writing with mkdir(parents=True, exist_ok=True).
# Use f-strings to format numeric output to a consistent number of decimal places.


def load_stats_report(*, stats: dict[str, float], out_path: Path) -> None:
    """L: Write stats to a text file in data/processed.

    Arguments:
        stats: Dictionary with statistics to write.
        out_path: Path to output text file.

    Returns:
        None
    """
    out_path.parent.mkdir(parents=True, exist_ok=True)

    with out_path.open("w", encoding="utf-8") as f:
        f.write("Texas Grade 8 NAEP Score Statistics By ELL Status (1990 - 2024)\n")
        if "Non" in out_path.stem:
            f.write("Non-ELL Stats:\n")
        else:
            f.write("ELL Stats:\n")
        f.write(f"Count: {int(stats['count'])}\n")
        f.write(f"Minimum: {stats['min']:.2f}\n")
        f.write(f"Maximum: {stats['max']:.2f}\n")
        f.write(f"Mean: {stats['mean']:.2f}\n")
        f.write(f"Median: {stats['median']:.2f}\n")
        f.write(f"Standard Deviation: {stats['stdev']:.2f}\n")


# === FULL PIPELINE ===

# This function composes the four steps into a single callable pipeline.
# The logger is passed in as an argument so this function works in any context.


def run_xlsx_pipeline_ELL(*, raw_dir: Path, processed_dir: Path, logger: Any) -> None:
    """Run the full ETVL pipeline.

    Arguments:
        raw_dir: Path to data/raw directory.
        processed_dir: Path to data/processed directory.
        logger: Logger for logging messages.

    Returns:
        None
    """
    logger.info("XLSX: START")

    input_file = raw_dir / "grade_8_ELL_math_scores_TX.xlsx"
    output_file_ELL = processed_dir / "grade_8_ELL_math_scores_TX_stats.txt"
    output_file_Non_ELL = processed_dir / "grade_8_Non_ELL_math_scores_TX_stats.txt"
    graph_file = processed_dir / "grade_8_ELL_math_scores_TX_graph.png"

    column_letter_scores = "D"
    column_letter_years = "A"
    column_letter_key = "C"

    logger.info(f"XLSX: USING input file: {input_file}, column: {column_letter_scores}")

    # E: Read numeric values from column D and years from column A.
    ELL_values, Non_ELL_values = extract_xlsx_column_tuples(
        file_path=input_file,
        column_letter=column_letter_scores,
        column_letter_key=column_letter_key,
    )
    years = extract_xlsx_column_values(
        file_path=input_file, column_letter=column_letter_years
    )

    # T: Calculate statistics for the numeric values and create plot.
    # print(f"ELL Values: {ELL_values}") #Debugging
    # print(f"Non-ELL Values: {Non_ELL_values}") #Debugging

    ELL_stats = transform_scores_to_stats(scores=ELL_values)
    Non_ELL_stats = transform_scores_to_stats(scores=Non_ELL_values)
    plot_scores(
        ELL_scores=ELL_values,
        Non_ELL_scores=Non_ELL_values,
        years=years,
        out_path=graph_file,
    )

    # V: Verify the statistics before writing.
    verify_stats(stats=ELL_stats)
    verify_stats(stats=Non_ELL_stats)

    # L: Write results to disk.
    load_stats_report(
        stats=ELL_stats,
        out_path=output_file_ELL.with_name("grade_8_ELL_math_scores_TX_stats.txt"),
    )
    load_stats_report(
        stats=Non_ELL_stats,
        out_path=output_file_Non_ELL.with_name(
            "grade_8_Non_ELL_math_scores_TX_stats.txt"
        ),
    )

    logger.info("XLSX: wrote %s", output_file_ELL)
    logger.info("XLSX: wrote %s", output_file_Non_ELL)
    logger.info("XLSX: created graph %s", graph_file)
    logger.info("XLSX: END")
